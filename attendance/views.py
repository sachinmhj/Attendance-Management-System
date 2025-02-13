from django.shortcuts import render, redirect, HttpResponse
from users.models import CustomUser
from .models import Attendance
from django.utils import timezone
from django.contrib import messages
# for excel
import openpyxl
from io import BytesIO
from datetime import datetime
# for pdf
from xhtml2pdf import pisa
from django.template.loader import render_to_string

# Create your views here.
# ---------- for student users ----------
def myattendance(request, sg):
    if request.user.is_authenticated and request.user.role == 'student':
        student = CustomUser.objects.get(user_slug=sg)
        
        # -- for downloading pdf --
        if request.method == "POST":
            template_path = render_to_string('attendance/student/downloadpdf.html', {"student":student})
            # Create the PDF from HTML content
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="student_list.pdf"'
            pisa_status = pisa.CreatePDF(template_path, dest=response)
            # If there was an error generating the PDF
            if pisa_status.err:
                return HttpResponse('Error generating PDF')
            return response
        # -- for downloading pdf --
        
        return render(request, 'attendance/student/attendancedata.html', {"student":student})
    else:
        return redirect('/')

def teachersassgined(request):
    if request.user.is_authenticated and request.user.role == 'student':
        assignedteach = CustomUser.objects.filter(assigned_classes=request.user.student_class)
        return render(request, 'attendance/student/teachersassigned.html', {"assignedteach": assignedteach})
    else:
        return redirect('/')
# ---------- for student users ----------

# ---------- for teacher users ----------
def allstudent(request):
    if request.user.is_authenticated and request.user.role == 'teacher':
        allgradelevel = request.user.assigned_classes.all()
        # print(allgradelevel)
        gradelevellist = []
        for i in allgradelevel:
            gradetitle = f'Grade {i}'
            gradestudents = CustomUser.objects.filter(student_class=i.grade_level)
            gradedetails = {'gradetitle': gradetitle, 'gradestudents': gradestudents}
            gradelevellist.append(gradedetails)
        # print(gradelevellist)
        return render(request, 'attendance/teacher/allstudents.html', {'gradelist': gradelevellist})
    else:
        return redirect('/')

def studentdetails(request, sg):
    if request.user.is_authenticated and request.user.role == 'teacher':
        stu = CustomUser.objects.get(user_slug=sg)
        return render(request, 'attendance/teacher/studentdetails.html', {'stu':stu})
    else:
        return redirect('/')

def selectgradeattend(request):
    if request.user.is_authenticated and request.user.role == 'teacher':
        assigned_grades = request.user.assigned_classes.all()
        # print(assigned_grades)
        return render(request, 'attendance/teacher/selectgradeattend.html', {'assigned_grades': assigned_grades})
    else:
        return redirect('/')

def takeattendance(request, gradelevel):
    if request.user.is_authenticated and request.user.role == 'teacher':
        students = CustomUser.objects.filter(student_class=gradelevel)
        if request.method == "POST":
            # check if attendance taken for today or not
            todaysdate = (timezone.now()).date()
            # attendancedatetime = Attendance.objects.values('todaydate', 'student_grade_level')
            # attendancedateonly = [x['todaydate'].date() for x in attendancedatetime]
            # attendancegrade = [x['student_grade_level'] for x in attendancedatetime]
            for stu in students:
                temp = list()
                allattendances = stu.user_attendance.all()
                if allattendances:
                    for userattendance in allattendances:
                        if(userattendance.todaydate.date() == todaysdate) and (userattendance.student_grade_level == gradelevel):
                            temp.append(True)
                        else:
                            temp.append(False)
                    # print(temp)
                    if True in temp:
                        messages.warning(request, 'Attendance has already been taken for today.')
                        break
            else:
                # print('inside else')
                # if attendance not taken for today
                for stu in students:
                    inputname = f'{stu.username}-{stu.id}'
                    checkbox_val = request.POST.get(inputname)
                    if checkbox_val:
                        saveattend = Attendance(user=stu, todaydate=timezone.now(), is_present=True, taken_by=request.user, student_grade_level=gradelevel)
                    else:
                        saveattend = Attendance(user=stu, todaydate=timezone.now(), is_present=False, taken_by=request.user, student_grade_level=gradelevel)
                    saveattend.save()
                messages.success(request, 'Attendance taken successfully.')
        return render(request, 'attendance/teacher/takeattendance.html', {"students": students})
    else:
        return redirect('/')

def selectattendancehistory(request):
    if request.user.is_authenticated and request.user.role == 'teacher':
        assigned_grades = request.user.assigned_classes.all()
        # print(assigned_grades)
        return render(request, 'attendance/teacher/selectattendhistory.html', {'assigned_grades': assigned_grades})
    else:
        return redirect('/')

def studentsattendancehistory(request, gradelevel):
    if request.user.is_authenticated and request.user.role == 'teacher':
        allstudents = CustomUser.objects.filter(student_class=gradelevel)
        
        # Show or Download specific date data
        if request.method == 'POST':
            fulldate = request.POST.get('date')
            if fulldate:
                specificdate = []
                fulldate_obj = datetime.strptime(fulldate, '%Y-%m-%d').date()
                allattendances = [x.user_attendance.all() for x in allstudents]
                for x in allattendances:
                    for y in x:
                        if y.todaydate.date() == fulldate_obj:
                            specificdate.append(y)
                if specificdate:
                    # -- for downloading excel ---
                    if 'downloadxl' in request.POST:
                        # print('start download')
                        wb = openpyxl.Workbook()
                        sheet1 = wb.active
                        sheet1.merge_cells('A1:E2')
                        sheet1['A1'] = f'Student Attendance Data - {fulldate}'
                        sheet1['A1'].font = openpyxl.styles.Font(bold=True)
                        sheet1['A1'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                        sheet1['A1'].fill = openpyxl.styles.PatternFill(start_color="FF6347", end_color="FFF000", fill_type="solid")
                        sheet1['A3']='Name'
                        sheet1['B3']='Grade'
                        sheet1['C3']='Date'
                        sheet1['D3']='Attendance Status'
                        sheet1['E3']='taken By'
                        for count, student in enumerate(specificdate):
                            sheet1.cell(row=count+4, column=1, value=student.user.username)
                            sheet1.cell(row=count+4, column=2, value=student.student_grade_level)
                            sheet1.cell(row=count+4, column=3, value=student.todaydate.date())
                            sheet1.cell(row=count+4, column=4, value='present' if(student.is_present) else 'absent')
                            sheet1.cell(row=count+4, column=5, value=student.taken_by.username)
                            
                        # -- Taken from stactoverflow --
                        # Write the Excel file to an in-memory buffer
                        excel_file = BytesIO()
                        wb.save(excel_file)
                        excel_file.seek(0)  # Go to the beginning of the file in memory
                        # Prepare the HTTP response to send the file to the user
                        response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = f'attachment; filename=StudentInfoForDate-{fulldate}.xlsx'
                        return response
                        # -- Taken from stactoverflow --
                        
                    # -- for downloading excel ---
                    else:
                        context = {'specificdate': specificdate, 'fulldate': fulldate}
                        return render(request, 'attendance/teacher/specificdaterecord.html', context)
                else:
                    messages.warning(request, 'Warning!!! No record found to download fo the given date.')
            else:
                messages.warning(request, 'Warning!!! Please select the date before submitting')
        
        # list1 = [x.user_attendance.all() for x in allstudents]
        return render(request, 'attendance/teacher/studentattendhistory.html',{"allstudents": allstudents})
    else:
        return redirect('/')
# ---------- for teacher users ----------